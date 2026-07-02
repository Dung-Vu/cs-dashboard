#!/usr/bin/env python3
"""CS Dashboard — KPI Dashboard for Customer Service (Complaints + Warranty + CSAT + SLA)."""

from __future__ import annotations

import calendar
import json
import os
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# WeasyPrint is imported lazily inside the PDF route so the rest of the app
# (Excel export, dashboard, JSON API) doesn't pay the ~150ms import cost on
# every cold start.
# from weasyprint import HTML

BASE_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(Path(__file__).parent))
from odoo_core import connect_rpc, search_read, search_count

ENV_PATH = BASE_DIR / ".env"
TEMPLATES_DIR = str(BASE_DIR / "templates")
app = Flask(__name__, template_folder=TEMPLATES_DIR)

# ── Helpers ──────────────────────────────────────────────────────────────────

def load_env():
    if ENV_PATH.exists():
        with open(ENV_PATH) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip().strip('"').strip("'")


_odoo_conn = None

def get_odoo():
    global _odoo_conn
    if _odoo_conn is None:
        load_env()
        _odoo_conn = connect_rpc("prod")
    return _odoo_conn


_cache: dict[str, tuple[float, Any]] = {}
CACHE_TTL = 300  # 5 minutes


def cached(key: str):
    """Cache decorator for KPI functions."""
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            cache_key = f"{key}:{json.dumps(args)}:{json.dumps(kwargs, sort_keys=True)}"
            now = time.time()
            if cache_key in _cache:
                ts, val = _cache[cache_key]
                if now - ts < CACHE_TTL:
                    return val
            result = fn(*args, **kwargs)
            _cache[cache_key] = (now, result)
            return result
        return wrapper
    return decorator


# ── Data Fetching ────────────────────────────────────────────────────────────

TEAM_IDS = [14, 15]  # ORD=14, BON=15
TEAM_TO_BRAND = {14: "ORD", 15: "BON"}

# CSAT thresholds (used only when real rating_last_value exists; not used here)
RATING_SATISFIED = 8  # >= 8 = satisfied
RATING_UNSATISFIED = 5  # <= 5 = unsatisfied

# CSAT feedback tag IDs. The Bonario/Ordinaire helpdesk encodes customer
# satisfaction as tags on the ticket (rating_last_value is always 0).
# We resolve these by NAME on first use, then cache — IDs differ across DBs.
_FEEDBACK_TAGS: dict | None = None


def _load_feedback_tags(conn) -> dict:
    """Resolve the helpdesk.tag IDs that represent the three CSAT outcomes.

    Returns dict with keys: 'satisfied', 'neutral', 'unsatisfied', 'complaint'.
    Any key may be None if the tag is missing in this database.
    """
    global _FEEDBACK_TAGS
    if _FEEDBACK_TAGS is not None:
        return _FEEDBACK_TAGS
    tags = search_read(conn, "helpdesk.tag", [], ["id", "name"])
    resolved = {"satisfied": None, "neutral": None, "unsatisfied": None, "complaint": None}
    for t in tags:
        name = (t.get("name") or "").strip().lower()
        if name == "feedback hài lòng":
            resolved["satisfied"] = t["id"]
        elif name == "feedback không hài lòng":
            resolved["unsatisfied"] = t["id"]
        elif name == "feedback kh":
            resolved["neutral"] = t["id"]
        elif name == "khiếu nại":
            resolved["complaint"] = t["id"]
    _FEEDBACK_TAGS = resolved
    return resolved


def feedback_score(ticket: dict) -> int | None:
    """Synthetic CSAT score (0-10) derived from feedback tags on a ticket.

    Priority when multiple feedback tags are present: satisfied > neutral > unsatisfied
    (treat a positive signal as overriding a negative one on the same ticket).
    Returns None if the ticket has no feedback tag.
    """
    tags = set(ticket.get("tag_ids") or [])
    f = _load_feedback_tags(get_odoo())
    if f["satisfied"] in tags:
        return 10
    if f["neutral"] in tags:
        return 7
    if f["unsatisfied"] in tags:
        return 3
    return None


def brand_of(team_id_field) -> str:
    """Map helpdesk team_id (list or int) to brand label."""
    if isinstance(team_id_field, list) and team_id_field:
        return TEAM_TO_BRAND.get(team_id_field[0], "other")
    if isinstance(team_id_field, int):
        return TEAM_TO_BRAND.get(team_id_field, "other")
    return "other"


def get_partner_brands(conn, partner_ids: list[int]) -> dict[int, str]:
    """Map partner_id -> brand based on most recent GH-KH task project_id."""
    if not partner_ids:
        return {}
    tasks = search_read(conn, "project.task", [
        ("name", "ilike", "GH-KH"),
        ("partner_id", "in", partner_ids),
        ("date_deadline", "!=", False),
    ], ["partner_id", "date_deadline", "project_id"])

    latest: dict[int, dict] = {}
    for t in tasks:
        p = t.get("partner_id")
        pid = p[0] if isinstance(p, list) and p else p
        if pid and (pid not in latest or (t.get("date_deadline") or "") > (latest[pid].get("date_deadline") or "")):
            latest[pid] = t

    brands: dict[int, str] = {}
    for pid in partner_ids:
        task = latest.get(pid, {})
        proj = task.get("project_id")
        proj_id = proj[0] if isinstance(proj, list) and proj else proj
        if proj_id == 10:
            brands[pid] = "BON"
        elif proj_id == 949:
            brands[pid] = "ORD"
        else:
            brands[pid] = "other"
    return brands


def partner_id_of(partner_field) -> int | None:
    """Extract id from partner_id field which may be [id, name] or scalar."""
    if isinstance(partner_field, list) and partner_field:
        return partner_field[0]
    return partner_field if isinstance(partner_field, int) else None


def fetch_tickets(team_ids: list[int] | None = None, date_from: str | None = None,
                  date_to: str | None = None) -> list[dict]:
    domain: list = [("team_id", "in", team_ids or TEAM_IDS)]
    if date_from:
        domain.append(("create_date", ">=", date_from))
    if date_to:
        domain.append(("create_date", "<=", date_to))
    conn = get_odoo()
    return search_read(conn, "helpdesk.ticket", domain, [
        "id", "name", "team_id", "stage_id", "user_id", "partner_id",
        "create_date", "write_date", "close_date",
        "x_studio_sale_orders", "x_studio_l_do_khiu_ni",
        "x_studio_product_ids", "x_studio_so",
        "tag_ids",
        "rating_last_value", "sla_success", "sla_fail",
        "sla_deadline",
    ])


def fetch_sale_orders(partner_ids: list[int], after_date: str) -> list[dict]:
    if not partner_ids:
        return []
    conn = get_odoo()
    return search_read(conn, "sale.order", [
        ("partner_id", "in", partner_ids),
        ("date_order", ">=", after_date),
        ("state", "not in", ["cancel"]),
    ], ["id", "partner_id", "date_order", "state", "name"])


# ── KPI Calculation ──────────────────────────────────────────────────────────

def classify_ticket(name: str) -> str:
    """Classify ticket type from name prefix."""
    if not name:
        return "Other"
    prefix = name.split(" / ")[0].strip() if " / " in name else name[:5]
    if "KN" in prefix:
        return "complaint"
    if "BH" in prefix:
        return "warranty"
    if "FB" in prefix:
        return "feedback"
    return "other"


def compute_complaint_kpis(tickets: list[dict], date_from: str, date_to: str):
    """FCR, Repeat Complaint Rate, Top 5 reasons (per brand), monthly case count."""
    complaints = [t for t in tickets if classify_ticket(t.get("name", "")) == "complaint"]

    # Group by BG code
    bg_groups: dict[str, list] = {}
    for t in complaints:
        bg = t.get("x_studio_sale_orders", "") or ""
        if bg:
            bg_groups.setdefault(bg, []).append(t)

    total = len(complaints)
    fcr_count = sum(1 for bg, items in bg_groups.items() if len(items) == 1)
    repeat_count = sum(1 for bg, items in bg_groups.items() if len(items) > 1)
    repeat_tickets = sum(len(items) for bg, items in bg_groups.items() if len(items) > 1)

    fcr_rate = round(fcr_count / max(len(bg_groups), 1) * 100, 1)
    repeat_rate = round(repeat_tickets / max(total, 1) * 100, 1)

    # Top 5 reasons — overall + split by brand
    overall_reasons = Counter()
    reasons_by_brand: dict[str, Counter] = {"BON": Counter(), "ORD": Counter()}
    for t in complaints:
        r = t.get("x_studio_l_do_khiu_ni", "") or "Không xác định"
        overall_reasons[r] += 1
        b = brand_of(t.get("team_id"))
        if b in reasons_by_brand:
            reasons_by_brand[b][r] += 1

    # Monthly trend (last 12 months)
    monthly = {}
    monthly_by_brand: dict[str, dict[str, int]] = {"BON": {}, "ORD": {}}
    current = datetime.strptime(date_to[:7] + "-01", "%Y-%m-%d")
    for i in range(12):
        m = (current - timedelta(days=30 * i)).strftime("%Y-%m")
        monthly[m] = 0
        monthly_by_brand["BON"][m] = 0
        monthly_by_brand["ORD"][m] = 0
    for t in complaints:
        m = (t.get("create_date", "") or "")[:7]
        if m in monthly:
            monthly[m] += 1
        b = brand_of(t.get("team_id"))
        if b in monthly_by_brand and m in monthly_by_brand[b]:
            monthly_by_brand[b][m] += 1
    monthly_sorted = sorted(monthly.items())

    # Compare this month vs last month
    this_month = sum(1 for t in complaints if (t.get("create_date", "") or "")[:7] == date_to[:7])
    last_month = sum(1 for t in complaints if (t.get("create_date", "") or "")[:7] == date_from[:7])

    return {
        "total": total,
        "fcr_rate": fcr_rate,
        "fcr_count": fcr_count,
        "repeat_rate": repeat_rate,
        "repeat_tickets": repeat_tickets,
        "top5_reasons": [{"reason": r, "count": c} for r, c in overall_reasons.most_common(5)],
        "top5_reasons_by_brand": {
            b: [{"reason": r, "count": c} for r, c in ctr.most_common(5)]
            for b, ctr in reasons_by_brand.items()
        },
        "monthly_trend": [{"month": m, "count": c} for m, c in monthly_sorted],
        "monthly_trend_by_brand": {
            b: [{"month": m, "count": monthly_by_brand[b][m]} for m, _ in monthly_sorted]
            for b in ("BON", "ORD")
        },
        "this_month": this_month,
        "last_month": last_month,
        "mom_change": round((this_month - last_month) / max(last_month, 1) * 100, 1) if last_month > 0 else 0,
    }


def compute_retention_rate(tickets: list[dict]):
    """Customers who bought again after their last complaint."""
    complaints = [t for t in tickets if classify_ticket(t.get("name", "")) == "complaint"]
    if not complaints:
        return {"rate": 0, "retained": 0, "total_complainers": 0}

    # Group by partner
    partner_last_complaint: dict[int, str] = {}
    for t in complaints:
        p = t.get("partner_id")
        pid = p[0] if isinstance(p, list) and p else None
        if pid:
            cd = t.get("create_date", "")
            if pid not in partner_last_complaint or cd > partner_last_complaint[pid]:
                partner_last_complaint[pid] = cd

    # Check sale orders after complaint date
    retained = set()
    for pid, last_cd in partner_last_complaint.items():
        orders = fetch_sale_orders([pid], last_cd)
        if orders:
            retained.add(pid)

    total_complainers = len(partner_last_complaint)
    return {
        "rate": round(len(retained) / max(total_complainers, 1) * 100, 1),
        "retained": len(retained),
        "total_complainers": total_complainers,
    }


def compute_warranty_kpis(tickets: list[dict], date_from: str, date_to: str):
    """First Time Fix Rate, Repeat Repair Rate, top 5 reasons (per brand), monthly count."""
    warranties = [t for t in tickets if classify_ticket(t.get("name", "")) == "warranty"]

    # Group by BG code
    bg_groups: dict[str, list] = {}
    for t in warranties:
        bg = t.get("x_studio_sale_orders", "") or ""
        if bg:
            bg_groups.setdefault(bg, []).append(t)

    total = len(warranties)
    first_fix = sum(1 for bg, items in bg_groups.items() if len(items) == 1)
    repeat_count = sum(1 for bg, items in bg_groups.items() if len(items) > 1)
    repeat_tickets = sum(len(items) for bg, items in bg_groups.items() if len(items) > 1)

    first_fix_rate = round(first_fix / max(len(bg_groups), 1) * 100, 1)
    repeat_repair_rate = round(repeat_tickets / max(total, 1) * 100, 1)

    # Top 5 reasons — overall + per brand
    overall_reasons = Counter()
    reasons_by_brand: dict[str, Counter] = {"BON": Counter(), "ORD": Counter()}
    for t in warranties:
        r = t.get("x_studio_l_do_khiu_ni", "") or "Không xác định"
        overall_reasons[r] += 1
        b = brand_of(t.get("team_id"))
        if b in reasons_by_brand:
            reasons_by_brand[b][r] += 1

    # Monthly trend
    monthly = {}
    monthly_by_brand: dict[str, dict[str, int]] = {"BON": {}, "ORD": {}}
    current = datetime.strptime(date_to[:7] + "-01", "%Y-%m-%d")
    for i in range(12):
        m = (current - timedelta(days=30 * i)).strftime("%Y-%m")
        monthly[m] = 0
        monthly_by_brand["BON"][m] = 0
        monthly_by_brand["ORD"][m] = 0
    for t in warranties:
        m = (t.get("create_date", "") or "")[:7]
        if m in monthly:
            monthly[m] += 1
        b = brand_of(t.get("team_id"))
        if b in monthly_by_brand and m in monthly_by_brand[b]:
            monthly_by_brand[b][m] += 1

    monthly_sorted = sorted(monthly.items())
    this_month = sum(1 for t in warranties if (t.get("create_date", "") or "")[:7] == date_to[:7])
    last_month = sum(1 for t in warranties if (t.get("create_date", "") or "")[:7] == date_from[:7])

    return {
        "total": total,
        "first_fix_rate": first_fix_rate,
        "first_fix_count": first_fix,
        "repeat_repair_rate": repeat_repair_rate,
        "repeat_repair_tickets": repeat_tickets,
        "top5_reasons": [{"reason": r, "count": c} for r, c in overall_reasons.most_common(5)],
        "top5_reasons_by_brand": {
            b: [{"reason": r, "count": c} for r, c in ctr.most_common(5)]
            for b, ctr in reasons_by_brand.items()
        },
        "monthly_trend": [{"month": m, "count": c} for m, c in monthly_sorted],
        "monthly_trend_by_brand": {
            b: [{"month": m, "count": monthly_by_brand[b][m]} for m, _ in monthly_sorted]
            for b in ("BON", "ORD")
        },
        "this_month": this_month,
        "last_month": last_month,
        "mom_change": round((this_month - last_month) / max(last_month, 1) * 100, 1) if last_month > 0 else 0,
    }


def compute_csat(tickets: list[dict]):
    """CSAT score synthesized from feedback tags (rating_last_value is always 0 here)."""
    # Warm the tag-id cache so the per-ticket helper is just a set lookup.
    _load_feedback_tags(get_odoo())
    total = len(tickets)
    scores = [s for s in (feedback_score(t) for t in tickets) if s is not None]
    rated = len(scores)
    if rated == 0:
        return {"score": 0, "rated": 0, "total": total, "response_rate": 0.0}
    avg = sum(scores) / rated
    return {
        "score": round(avg, 1),
        "rated": rated,
        "total": total,
        "response_rate": round(rated / max(total, 1) * 100, 1),
    }


def _brand_month_buckets(tickets: list[dict], date_to: str, months_back: int = 12):
    """Initialize per-brand per-month buckets covering the last N months.

    'rated' holds synthetic CSAT scores derived from feedback tags
    (rating_last_value is always 0 in this database).
    'perfect' counts tickets tagged as 'Feedback hài lòng' (the 10/10 bucket).
    """
    months: list[str] = []
    current = datetime.strptime(date_to[:7] + "-01", "%Y-%m-%d")
    for i in range(months_back):
        months.append((current - timedelta(days=30 * i)).strftime("%Y-%m"))
    months.reverse()

    total = {b: {m: 0 for m in months} for b in ("BON", "ORD")}
    rated = {b: {m: [] for m in months} for b in ("BON", "ORD")}
    perfect = {b: {m: 0 for m in months} for b in ("BON", "ORD")}
    with_order = {b: {m: 0 for m in months} for b in ("BON", "ORD")}

    # Pre-resolve tag IDs once per call (cached after first call).
    f = _load_feedback_tags(get_odoo())
    satisfied_id = f["satisfied"]

    for t in tickets:
        b = brand_of(t.get("team_id"))
        if b not in total:
            continue
        m = (t.get("create_date", "") or "")[:7]
        if m not in total[b]:
            continue
        total[b][m] += 1
        score = feedback_score(t)
        if score is not None:
            rated[b][m].append(score)
            if score == 10 and satisfied_id is not None and satisfied_id in (t.get("tag_ids") or []):
                perfect[b][m] += 1
        if t.get("x_studio_sale_orders"):
            with_order[b][m] += 1

    return months, total, rated, perfect, with_order


def compute_csat_by_brand(tickets: list[dict], date_to: str, months_back: int = 12):
    """CSAT breakdown by brand with monthly trend (perfect_rate = % rating == 10)."""
    months, total, rated, perfect, _ = _brand_month_buckets(tickets, date_to, months_back)

    def build(rated_dict, total_dict, perfect_dict):
        all_ratings = [r for ms in rated_dict.values() for r in ms]
        all_total = sum(total_dict.values())
        rated_count = len(all_ratings)
        perfect_count = sum(perfect_dict.values())
        monthly = []
        for m in months:
            rs = rated_dict[m]
            monthly.append({
                "month": m,
                "score": round(sum(rs) / len(rs), 1) if rs else 0,
                "rated": len(rs),
                "total": total_dict[m],
                "perfect_rate": round(perfect_dict[m] / max(len(rs), 1) * 100, 1),
            })
        return {
            "score": round(sum(all_ratings) / rated_count, 1) if rated_count else 0,
            "rated": rated_count,
            "total": all_total,
            "response_rate": round(rated_count / max(all_total, 1) * 100, 1),
            "perfect_rate": round(perfect_count / max(rated_count, 1) * 100, 1),
            "monthly": monthly,
        }

    return {
        "months": months,
        "BON": build(rated["BON"], total["BON"], perfect["BON"]),
        "ORD": build(rated["ORD"], total["ORD"], perfect["ORD"]),
    }


def compute_satisfaction_split(tickets: list[dict]):
    """Count satisfied / neutral / unsatisfied tickets per brand using feedback tags."""
    _load_feedback_tags(get_odoo())
    result = {
        b: {"satisfied": 0, "neutral": 0, "unsatisfied": 0, "total_rated": 0}
        for b in ("BON", "ORD")
    }
    for t in tickets:
        tags = set(t.get("tag_ids") or [])
        if not tags:
            continue
        f = _load_feedback_tags(get_odoo())
        # Classify by which feedback tag is present (satisfied wins on ties).
        bucket = None
        if f["satisfied"] in tags:
            bucket = "satisfied"
        elif f["neutral"] in tags:
            bucket = "neutral"
        elif f["unsatisfied"] in tags:
            bucket = "unsatisfied"
        if bucket is None:
            continue
        b = brand_of(t.get("team_id"))
        if b not in result:
            continue
        result[b]["total_rated"] += 1
        result[b][bucket] += 1
    return result


def compute_response_and_order_rate(tickets: list[dict], date_to: str, months_back: int = 12):
    """Response rate (rated/total) and order rate (tickets-with-order/total) per month per brand."""
    months, total, rated, _, with_order = _brand_month_buckets(tickets, date_to, months_back)

    def build(total_dict, rated_dict, order_dict):
        return [
            {
                "month": m,
                "total": total_dict[m],
                "rated": len(rated_dict[m]),
                "response_rate": round(len(rated_dict[m]) / max(total_dict[m], 1) * 100, 1),
                "order_rate": round(order_dict[m] / max(total_dict[m], 1) * 100, 1),
            }
            for m in months
        ]

    return {
        "months": months,
        "BON": build(total["BON"], rated["BON"], with_order["BON"]),
        "ORD": build(total["ORD"], rated["ORD"], with_order["ORD"]),
    }


def compute_sla(tickets: list[dict]):
    """SLA completion rate by user and team.

    Only counts users whose employee is in a Customer Service (CSKH) department.
    Non-CS assignees (Sales, BIS, CEO office, ...) are excluded so the table
    reflects the CS team's actual SLA performance, not cross-team noise.
    """
    cs_user_ids = _cs_user_ids(get_odoo())
    users: dict[int, dict] = {}

    for t in tickets:
        user_id_field = t.get("user_id")
        if isinstance(user_id_field, list) and len(user_id_field) >= 2:
            uid = user_id_field[0]
            uname = user_id_field[1]
        else:
            uid = user_id_field[0] if isinstance(user_id_field, list) else user_id_field
            uname = str(uid)
        if not uid:
            continue
        if cs_user_ids and uid not in cs_user_ids:
            continue  # Skip non-CS assignees (e.g. Sales, BIS, CEO office)

        if uid not in users:
            users[uid] = {"name": uname, "total": 0, "success": 0, "fail": 0}

        users[uid]["total"] += 1
        if t.get("sla_success"):
            users[uid]["success"] += 1
        if t.get("sla_fail"):
            users[uid]["fail"] += 1

    team_total = sum(u["total"] for u in users.values())
    team_success = sum(u["success"] for u in users.values())

    return {
        "team_rate": round(team_success / max(team_total, 1) * 100, 1) if team_total > 0 else 0,
        "team_total": team_total,
        "team_success": team_success,
        "users": [
            {
                "name": u["name"],
                "rate": round(u["success"] / max(u["total"], 1) * 100, 1),
                "total": u["total"],
                "success": u["success"],
            }
            for u in sorted(users.values(), key=lambda x: x["total"], reverse=True)
        ],
    }


# ── CS department membership ──────────────────────────────────────────────────
# CSKH = Customer Service team. Their tickets are the ones that should be measured
# in the SLA report. Resolved by department name (leaf segment after last "/").
_CS_DEPT_IDS: set[int] | None = None
_CS_USER_IDS: set[int] | None = None


def _load_cs_department_ids(conn) -> set[int]:
    """Department IDs whose leaf name identifies the CS / CSKH team."""
    global _CS_DEPT_IDS
    if _CS_DEPT_IDS is not None:
        return _CS_DEPT_IDS
    depts = search_read(conn, "hr.department", [], ["id", "name"])
    cs_ids: set[int] = set()
    for d in depts:
        name = (d.get("name") or "").strip()
        leaf = name.split("/")[-1].strip().lower()
        if leaf in ("customer services", "cskh", "chăm sóc khách hàng"):
            cs_ids.add(d["id"])
    _CS_DEPT_IDS = cs_ids
    return cs_ids


def _cs_user_ids(conn) -> set[int]:
    """res.users IDs whose employee is in a CS department.

    Returns an empty set if no CS department is defined (caller should treat
    that as "filter everything out" rather than "show everyone").
    """
    global _CS_USER_IDS
    if _CS_USER_IDS is not None:
        return _CS_USER_IDS
    dept_ids = _load_cs_department_ids(conn)
    if not dept_ids:
        _CS_USER_IDS = set()
        return _CS_USER_IDS
    emps = search_read(
        conn, "hr.employee",
        [("department_id", "in", list(dept_ids))],
        ["user_id"],
    )
    _CS_USER_IDS = {
        e["user_id"][0] for e in emps
        if e.get("user_id") and isinstance(e["user_id"], list)
    }
    return _CS_USER_IDS


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    """Main dashboard page."""
    return render_template("dashboard.html")


def resolve_date_range(mode: str, date_str: str, months: int):
    """Return (from_date, to_date, mode, date_str) based on filter params.

    mode='date'  → single calendar day (date_str = YYYY-MM-DD)
    mode='range' → trailing N months ending today (months = N)
    """
    if mode == "date" and date_str:
        try:
            y, m, d = date_str.split("-")
            y, m, d = int(y), int(m), int(d)
            from_date = f"{y:04d}-{m:02d}-{d:02d} 00:00:00"
            to_date = f"{y:04d}-{m:02d}-{d:02d} 23:59:59"
            return from_date, to_date, "date", date_str
        except (ValueError, TypeError):
            pass
    to_date = datetime.now().strftime("%Y-%m-%d 23:59:59")
    from_date = (datetime.now() - timedelta(days=30 * months)).strftime("%Y-%m-%d 00:00:00")
    return from_date, to_date, "range", ""


@app.route("/api/kpis")
def api_kpis():
    """JSON endpoint: all KPIs."""
    team = request.args.get("team", "all")  # all, ord, bon
    mode = request.args.get("mode", "range")  # range, month
    months = int(request.args.get("months", "1"))
    month = request.args.get("date", "")  # YYYY-MM

    from_date, to_date, mode, month = resolve_date_range(mode, month, months)

    # Determine team filter
    team_ids_map = {"ord": [14], "bon": [15], "all": TEAM_IDS}
    team_ids = team_ids_map.get(team, TEAM_IDS)

    start = time.time()
    tickets = fetch_tickets(team_ids, from_date, to_date)

    complaint = compute_complaint_kpis(tickets, from_date, to_date)
    retention = compute_retention_rate(tickets)
    warranty = compute_warranty_kpis(tickets, from_date, to_date)
    csat = compute_csat(tickets)
    csat_brand = compute_csat_by_brand(tickets, to_date)
    satisfaction = compute_satisfaction_split(tickets)
    rates = compute_response_and_order_rate(tickets, to_date)
    sla = compute_sla(tickets)

    # Fetch delivery orders for on-time + BD metrics
    delivery_orders = fetch_delivery_orders(from_date, to_date)

    elapsed = int((time.time() - start) * 1000)

    return jsonify({
        "team": team,
        "mode": mode,
        "month": month,
        "period": f"{from_date[:10]} → {to_date[:10]}",
        "elapsed_ms": elapsed,
        "complaints": complaint,
        "retention": retention,
        "warranty": warranty,
        "csat": csat,
        "csat_by_brand": csat_brand,
        "satisfaction": satisfaction,
        "rates": rates,
        "sla": sla,
        "purchase": compute_purchase_metrics(months),
        "delivery": compute_delivery_metrics(delivery_orders),
        "bd_progress": compute_bd_progress_metrics(delivery_orders),
    })


def fetch_delivery_orders(from_date: str, to_date: str) -> list[dict]:
    conn = get_odoo()
    orders = search_read(conn, "sale.order", [
        ("date_order", ">=", from_date),
        ("date_order", "<=", to_date),
        ("state", "in", ["sale", "done"]),
    ], ["id", "name", "partner_id", "x_studio_ngy_hon_thnh", "date_order"])

    partner_ids = list(set(
        o["partner_id"][0] if isinstance(o.get("partner_id"), list) and o.get("partner_id") else o.get("partner_id")
        for o in orders if o.get("partner_id")
    ))

    # Fetch GH-KH tasks for the same partners
    tasks = {}
    if partner_ids:
        gh_tasks = search_read(conn, "project.task", [
            ("name", "ilike", "GH-KH"),
            ("partner_id", "in", partner_ids),
            ("date_deadline", "!=", False),
        ], ["id", "name", "partner_id", "date_deadline", "project_id"])

        for t in gh_tasks:
            pid = t["partner_id"][0] if isinstance(t.get("partner_id"), list) else t.get("partner_id")
            if pid and (pid not in tasks or (t.get("date_deadline") or "") > (tasks[pid].get("date_deadline") or "")):
                tasks[pid] = t  # Keep the latest task per partner

    for o in orders:
        pid = o["partner_id"][0] if isinstance(o.get("partner_id"), list) and o.get("partner_id") else o.get("partner_id")
        o["_gh_task"] = tasks.get(pid, {})
        # Determine brand: BON=10, ORD=949
        proj_id = o["_gh_task"].get("project_id")
        proj_id = proj_id[0] if isinstance(proj_id, list) and proj_id else proj_id
        o["_brand"] = "BON" if proj_id == 10 else ("ORD" if proj_id == 949 else "other")

    return orders


def compute_delivery_metrics(orders: list[dict]):
    """On-time delivery: GH-KH task date_deadline vs sale.order BD tiến độ dự kiến, split by brand."""
    def calc(subset):
        total = len(subset)
        on_time = 0
        late = 0
        for o in subset:
            planned = (o.get("x_studio_ngy_hon_thnh", "") or "")[:10]
            task = o.get("_gh_task", {})
            actual = (task.get("date_deadline", "") or "")[:10]
            if planned and actual:
                if actual <= planned:
                    on_time += 1
                else:
                    late += 1
        return {
            "total": total,
            "with_dates": on_time + late,
            "on_time": on_time,
            "late": late,
            "rate": round(on_time / max(on_time + late, 1) * 100, 1),
        }

    bon = [o for o in orders if o.get("_brand") == "BON"]
    ord_ = [o for o in orders if o.get("_brand") == "ORD"]

    return {
        "bonario": calc(bon),
        "ordinaire": calc(ord_),
        "total": len(orders),
    }


def compute_bd_progress_metrics(orders: list[dict]):
    """BD Tiến độ dự kiến: planned vs actual completion."""
    with_dates = [o for o in orders
                  if (o.get("x_studio_ngy_hon_thnh") or "") and (o.get("x_studio_date_of_completion") or "")]
    total = len(with_dates)

    on_time = 0
    late = 0
    for o in with_dates:
        planned = str(o.get("x_studio_ngy_hon_thnh", ""))[:10]
        actual = str(o.get("x_studio_date_of_completion", ""))[:10]
        if actual <= planned:
            on_time += 1
        else:
            late += 1

    return {
        "total": total,
        "on_time": on_time,
        "late": late,
        "rate": round(on_time / max(total, 1) * 100, 1),
    }


def compute_purchase_metrics(months: int = 12):
    """Purchase behavior metrics from sale.order, split by brand."""
    conn = get_odoo()
    now = datetime.now()
    this_year = now.year
    to_date = now.strftime("%Y-%m-%d 23:59:59")

    # Fetch orders in selected period
    from_date = (now - timedelta(days=30 * months)).strftime("%Y-%m-%d 00:00:00")
    orders = search_read(conn, "sale.order", [
        ("date_order", ">=", from_date),
        ("date_order", "<=", to_date),
        ("state", "in", ["sale", "done"]),
    ], ["id", "partner_id", "date_order", "amount_total", "name"])

    # Fetch orders from previous year(s) for cross-year comparison
    prev_to = (datetime(this_year, 1, 1) - timedelta(days=1)).strftime("%Y-%m-%d 23:59:59")
    prev_from = (datetime(this_year - 1, 1, 1)).strftime("%Y-%m-%d 00:00:00")
    prev_orders = search_read(conn, "sale.order", [
        ("date_order", ">=", prev_from),
        ("date_order", "<=", prev_to),
        ("state", "in", ["sale", "done"]),
    ], ["id", "partner_id", "date_order", "amount_total"])

    # Resolve brand per partner (using most recent GH-KH task)
    all_partner_ids = list({
        partner_id_of(o.get("partner_id"))
        for o in orders + prev_orders
        if partner_id_of(o.get("partner_id"))
    })
    partner_brands = get_partner_brands(conn, all_partner_ids)

    def brand_of_order(o):
        return partner_brands.get(partner_id_of(o.get("partner_id")) or -1, "other")

    def compute_brand_block(brand_orders, brand_prev_orders):
        prev_partners = set()
        for o in brand_prev_orders:
            pid = partner_id_of(o.get("partner_id"))
            if pid:
                prev_partners.add(pid)

        this_year_partners = set()
        partner_orders_this_year: dict[int, list] = {}
        for o in brand_orders:
            pid = partner_id_of(o.get("partner_id"))
            if not pid:
                continue
            if (o.get("date_order", "") or "")[:4] == str(this_year):
                this_year_partners.add(pid)
                partner_orders_this_year.setdefault(pid, []).append(o)

        cross_year = this_year_partners & prev_partners
        same_year_repeat = sum(1 for ol in partner_orders_this_year.values() if len(ol) >= 2)
        amounts = [o["amount_total"] for o in brand_orders if o.get("amount_total")]
        avg_order_value = round(sum(amounts) / max(len(amounts), 1), 2)

        return {
            "cross_year_returnees": len(cross_year),
            "prev_year_buyers": len(prev_partners),
            "same_year_repeat": same_year_repeat,
            "this_year_buyers": len(this_year_partners),
            "avg_order_value": avg_order_value,
            "total_orders": len(brand_orders),
        }

    # Split orders by brand
    by_brand_orders = {"BON": [], "ORD": []}
    by_brand_prev = {"BON": [], "ORD": []}
    for o in orders:
        b = brand_of_order(o)
        if b in by_brand_orders:
            by_brand_orders[b].append(o)
    for o in prev_orders:
        b = brand_of_order(o)
        if b in by_brand_prev:
            by_brand_prev[b].append(o)

    overall = compute_brand_block(orders, prev_orders)
    by_brand = {
        b: compute_brand_block(by_brand_orders[b], by_brand_prev[b])
        for b in ("BON", "ORD")
    }

    # Monthly orders (overall, last 12 months)
    monthly_orders = {}
    for i in range(12):
        m = (now - timedelta(days=30 * i)).strftime("%Y-%m")
        monthly_orders[m] = 0
    for o in orders:
        m = (o.get("date_order", "") or "")[:7]
        if m in monthly_orders:
            monthly_orders[m] += 1

    monthly_by_brand: dict[str, dict[str, int]] = {b: dict(monthly_orders) for b in ("BON", "ORD")}
    for b in ("BON", "ORD"):
        for k in monthly_by_brand[b]:
            monthly_by_brand[b][k] = 0
    for o in orders:
        b = brand_of_order(o)
        if b in monthly_by_brand:
            m = (o.get("date_order", "") or "")[:7]
            if m in monthly_by_brand[b]:
                monthly_by_brand[b][m] += 1

    return {
        **overall,
        "by_brand": by_brand,
        "monthly_orders": [{"month": m, "count": c} for m, c in sorted(monthly_orders.items())],
        "monthly_orders_by_brand": {
            b: [{"month": m, "count": monthly_by_brand[b][m]} for m in sorted(monthly_orders)]
            for b in ("BON", "ORD")
        },
    }


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


# ── Excel export ───────────────────────────────────────────────────────────────
# Multi-sheet workbook — one sheet per dashboard section, plus a pivot-friendly
# long-format "Monthly trends" sheet.

_HEADER_FILL = PatternFill(start_color="0D6B62", end_color="0D6B62", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEAD_FILL = PatternFill(start_color="ECEDEA", end_color="ECEDEA", fill_type="solid")
_SUBHEAD_FONT = Font(bold=True, color="1A1814", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1A1814")
_TOTAL_FONT = Font(bold=True, color="1A1814")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN = Side(border_style="thin", color="D6D2C8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _autosize(ws, min_width: int = 12, max_width: int = 50) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                longest = max(longest, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(longest + 2, max_width))


def _write_header_row(ws, row_idx, headers):
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER


def _write_row(ws, row_idx, values, alignments=None, number_formats=None):
    for col_idx, v in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=v)
        c.border = _BORDER
        if alignments and col_idx - 1 < len(alignments):
            c.alignment = alignments[col_idx - 1]
        if number_formats and col_idx - 1 < len(number_formats) and number_formats[col_idx - 1]:
            c.number_format = number_formats[col_idx - 1]


def _add_overview_sheet(wb, data):
    ws = wb.create_sheet("Tong quan")
    ws["A1"] = "CS Dashboard — Bao cao tong quan"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = "Ky bao cao:"
    ws["B2"] = data.get("period", "")
    ws["A3"] = "Team:"
    team_label = {"all": "Tat ca team", "ord": "Khach Ordinaire", "bon": "Khach Bonario"}.get(
        data.get("team"), data.get("team"))
    ws["B3"] = team_label
    ws["A4"] = "Ngay xuat:"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    c = data.get("complaints", {})
    w = data.get("warranty", {})
    csat = data.get("csat", {})
    sla = data.get("sla", {})
    ret = data.get("retention", {})
    bd = data.get("bd_progress", {})

    _write_header_row(ws, 6, ["Phan", "Chi so", "Gia tri", "Don vi", "Ghi chu"])
    rows = [
        ("Khieu nai", "FCR Rate", c.get("fcr_rate", 0), "%", ">=80% la tot"),
        ("Khieu nai", "Repeat Rate", c.get("repeat_rate", 0), "%", "<=30% la tot"),
        ("Khieu nai", "Tong case", c.get("total", 0), "case", ""),
        ("Bao hanh", "First Fix Rate", w.get("first_fix_rate", 0), "%", ">=80% la tot"),
        ("Bao hanh", "Repeat Repair", w.get("repeat_repair_rate", 0), "%", "<=30% la tot"),
        ("Bao hanh", "Tong case", w.get("total", 0), "case", ""),
        ("CSAT", "Diem trung binh", csat.get("score", 0), "/10", ">=8 la tot"),
        ("CSAT", "Ty le danh gia", csat.get("response_rate", 0), "%", ""),
        ("CSAT", "Tong ticket", csat.get("total", 0), "ticket", ""),
        ("Retention", "Ty le mua lai", ret.get("rate", 0), "%", ""),
        ("SLA CSKH", "Team SLA", sla.get("team_rate", 0), "%",
         f"{sla.get('team_success', 0)}/{sla.get('team_total', 0)} dat"),
        ("BD Tien do", "Dung tien do", bd.get("rate", 0), "%",
         f"{bd.get('on_time', 0)}/{bd.get('total', 0)}"),
    ]
    for i, (section, label, val, unit, note) in enumerate(rows, start=7):
        _write_row(ws, i, [section, label, val, unit, note],
                   alignments=[_LEFT, _LEFT, _RIGHT, _LEFT, _LEFT])
        ws.cell(row=i, column=3).number_format = "0.0"
    _autosize(ws, min_width=14, max_width=42)


def _add_kv_sheet(wb, sheet_name, title, rows):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")
    _write_header_row(ws, 3, ["Chi so", "Gia tri", "Don vi", "Ghi chu"])
    for i, (label, val, unit, note) in enumerate(rows, start=4):
        _write_row(ws, i, [label, val, unit, note],
                   alignments=[_LEFT, _RIGHT, _LEFT, _LEFT])
        ws.cell(row=i, column=2).number_format = "0.0"
    _autosize(ws, min_width=18, max_width=42)


def _add_perbrand_reasons_sheet(wb, sheet_name, by_brand):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = sheet_name.replace("-", " — ")
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:C1")
    row = 3
    for brand, label in (("BON", "Bonario"), ("ORD", "Ordinaire")):
        ws.cell(row=row, column=1, value=f"--- {label} ---").font = _SUBHEAD_FONT
        ws.cell(row=row, column=1).fill = _SUBHEAD_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        _write_header_row(ws, row, ["Hang", "Nguyen nhan", "So case"])
        row += 1
        for j, r in enumerate(by_brand.get(brand, []), start=row):
            _write_row(ws, j, [j - row + 1, r["reason"], r["count"]],
                       alignments=[_CENTER, _LEFT, _RIGHT])
        row = ws.max_row + 2
    _autosize(ws, min_width=10, max_width=48)


def _add_brand_monthly_sheet(wb, sheet_name, by_brand, title):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")
    _write_header_row(ws, 3, ["Thang", "Tong case BON", "Tong case ORD", "Tong"])
    bon_m = {m["month"]: m["count"] for m in by_brand.get("BON", [])}
    ord_m = {m["month"]: m["count"] for m in by_brand.get("ORD", [])}
    months = sorted(set(bon_m) | set(ord_m))
    for i, m in enumerate(months, start=4):
        _write_row(ws, i, [m, bon_m.get(m, 0), ord_m.get(m, 0),
                           bon_m.get(m, 0) + ord_m.get(m, 0)],
                   alignments=[_CENTER, _RIGHT, _RIGHT, _RIGHT])
    _autosize(ws, min_width=12)


def _add_complaint_sheet(wb, data):
    c = data.get("complaints", {})
    rows = [
        ("FCR Rate", c.get("fcr_rate", 0), "%", ">=80% la tot"),
        ("FCR count", c.get("fcr_count", 0), "case", "Lan dau xu ly dung"),
        ("Repeat Rate", c.get("repeat_rate", 0), "%", "<=30% la tot"),
        ("Repeat tickets", c.get("repeat_tickets", 0), "case", ""),
        ("Tong case", c.get("total", 0), "case", ""),
        ("Thang nay", c.get("this_month", 0), "case", ""),
        ("Thang truoc", c.get("last_month", 0), "case", ""),
        ("MoM change", c.get("mom_change", 0), "%", "So voi thang truoc"),
    ]
    _add_kv_sheet(wb, "Khieu nai", "Khiếu nại — FCR & Repeat", rows)
    _add_perbrand_reasons_sheet(wb, "KN - Top nguyen nhan",
                                c.get("top5_reasons_by_brand", {}))
    _add_brand_monthly_sheet(wb, "KN - Xu huong thang",
                             c.get("monthly_trend_by_brand", {}),
                             "Khiếu nại — Xu hướng theo tháng")


def _add_warranty_sheet(wb, data):
    w = data.get("warranty", {})
    rows = [
        ("First Fix Rate", w.get("first_fix_rate", 0), "%", ">=80% la tot"),
        ("First Fix count", w.get("first_fix_count", 0), "case", ""),
        ("Repeat Repair Rate", w.get("repeat_repair_rate", 0), "%", "<=30% la tot"),
        ("Repeat Repair tickets", w.get("repeat_repair_tickets", 0), "case", ""),
        ("Tong case", w.get("total", 0), "case", ""),
        ("Thang nay", w.get("this_month", 0), "case", ""),
        ("Thang truoc", w.get("last_month", 0), "case", ""),
    ]
    _add_kv_sheet(wb, "Bao hanh", "Bảo hành — First Time Fix & Repeat Repair", rows)
    _add_perbrand_reasons_sheet(wb, "BH - Top nguyen nhan",
                                w.get("top5_reasons_by_brand", {}))
    _add_brand_monthly_sheet(wb, "BH - Xu huong thang",
                             w.get("monthly_trend_by_brand", {}),
                             "Bảo hành — Xu hướng theo tháng")


def _add_csat_sheets(wb, data):
    csat = data.get("csat", {})
    sat = data.get("satisfaction", {})
    csat_brand = data.get("csat_by_brand", {})

    rows = [
        ("Diem trung binh", csat.get("score", 0), "/10", ">=8 la tot"),
        ("Ty le danh gia", csat.get("response_rate", 0), "%", ""),
        ("Da danh gia", csat.get("rated", 0), "ticket", ""),
        ("Tong ticket", csat.get("total", 0), "ticket", ""),
    ]
    _add_kv_sheet(wb, "CSAT - Tong quan", "CSAT — Tổng quan", rows)

    ws = wb.create_sheet("CSAT - Theo Brand")
    ws["A1"] = "CSAT theo Brand"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")
    _write_header_row(ws, 3, ["Brand", "Score (/10)", "Response (%)",
                              "Perfect 10 (%)", "Rated / Total"])
    for i, brand in enumerate(("BON", "ORD"), start=4):
        b = csat_brand.get(brand, {})
        _write_row(ws, i, [
            "Bonario Vietnam" if brand == "BON" else "Ordinaire Vietnam",
            b.get("score", 0), b.get("response_rate", 0),
            b.get("perfect_rate", 0),
            f"{b.get('rated', 0)} / {b.get('total', 0)}",
        ], alignments=[_LEFT, _RIGHT, _RIGHT, _RIGHT, _CENTER],
           number_formats=["", "0.0", "0.0", "0.0", ""])
    _autosize(ws, min_width=18, max_width=28)

    ws = wb.create_sheet("CSAT - Phan bo")
    ws["A1"] = "Phan bo muc do hai long theo Brand"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")
    _write_header_row(ws, 3, ["Brand", "Hai long", "Trung binh",
                              "Khong hai long", "Tong rated"])
    for i, brand in enumerate(("BON", "ORD"), start=4):
        s = sat.get(brand, {})
        _write_row(ws, i, [
            "Bonario Vietnam" if brand == "BON" else "Ordinaire Vietnam",
            s.get("satisfied", 0), s.get("neutral", 0),
            s.get("unsatisfied", 0), s.get("total_rated", 0),
        ], alignments=[_LEFT, _RIGHT, _RIGHT, _RIGHT, _RIGHT])
    _autosize(ws, min_width=18)


def _add_purchase_sheets(wb, data):
    p = data.get("purchase", {})
    ret = data.get("retention", {})

    rows = [
        ("Ty le mua lai", ret.get("rate", 0), "%", ""),
        ("Khach mua lai", ret.get("retained", 0), "khach", ""),
        ("Tong khach khieu nai", ret.get("total_complainers", 0), "khach", ""),
    ]
    _add_kv_sheet(wb, "Mua hang - Retention",
                  "Khách hàng quay lại sau khiếu nại", rows)

    ws = wb.create_sheet("Mua hang - Theo Brand")
    ws["A1"] = "Hanh vi mua hang theo Brand"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:F1")
    _write_header_row(ws, 3, ["Brand", "Khach cu quay lai", "Mua lap cung nam",
                              "Tong buyer trong nam", "Gia tri DTB (VND)", "Tong don"])
    _write_row(ws, 4, [
        "Tong cong",
        p.get("cross_year_returnees", 0), p.get("same_year_repeat", 0),
        p.get("this_year_buyers", 0), p.get("avg_order_value", 0),
        p.get("total_orders", 0),
    ], alignments=[_LEFT, _RIGHT, _RIGHT, _RIGHT, _RIGHT, _RIGHT],
       number_formats=["", "0", "0", "0", "#,##0", "0"])
    for i, brand in enumerate(("BON", "ORD"), start=5):
        b = p.get("by_brand", {}).get(brand, {})
        _write_row(ws, i, [
            "Bonario Vietnam" if brand == "BON" else "Ordinaire Vietnam",
            b.get("cross_year_returnees", 0), b.get("same_year_repeat", 0),
            b.get("this_year_buyers", 0), b.get("avg_order_value", 0),
            b.get("total_orders", 0),
        ], alignments=[_LEFT, _RIGHT, _RIGHT, _RIGHT, _RIGHT, _RIGHT],
           number_formats=["", "0", "0", "0", "#,##0", "0"])
    _autosize(ws, min_width=18, max_width=26)


def _add_ops_sheets(wb, data):
    d = data.get("delivery", {})
    bd = data.get("bd_progress", {})
    sla = data.get("sla", {})

    ws = wb.create_sheet("GH - Giao hang dung hen")
    ws["A1"] = "Giao hang dung lich theo Brand"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")
    _write_header_row(ws, 3, ["Brand", "Dung lich (%)", "Dung han", "Tre han", "Tong SO"])
    for i, (brand, label) in enumerate(
            (("bonario", "Bonario Vietnam"), ("ordinaire", "Ordinaire Vietnam")), start=4):
        b = d.get(brand, {})
        _write_row(ws, i, [label, b.get("rate", 0), b.get("on_time", 0),
                           b.get("late", 0), b.get("total", 0)],
                   alignments=[_LEFT, _RIGHT, _RIGHT, _RIGHT, _RIGHT],
                   number_formats=["", "0.0", "0", "0", "0"])
    _autosize(ws, min_width=18)

    rows = [
        ("Dung tien do", bd.get("rate", 0), "%", f"{bd.get('on_time',0)}/{bd.get('total',0)}"),
        ("Dung han", bd.get("on_time", 0), "case", ""),
        ("Tre han", bd.get("late", 0), "case", ""),
        ("Tong", bd.get("total", 0), "case", ""),
    ]
    _add_kv_sheet(wb, "BD - Tien do", "BD — Tiến độ dự kiến", rows)

    ws = wb.create_sheet("SLA - Ca nhan")
    ws["A1"] = "SLA CSKH — Team & Ca nhan"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = (f"Team SLA: {sla.get('team_rate',0)}%  "
                f"({sla.get('team_success',0)} / {sla.get('team_total',0)})")
    ws["A2"].font = _TOTAL_FONT
    ws.merge_cells("A2:E2")
    _write_header_row(ws, 4, ["Nhan su", "SLA Rate (%)", "Tong ticket",
                              "Dat SLA", "That bai"])
    for i, u in enumerate(sla.get("users", []), start=5):
        _write_row(ws, i, [u["name"], u["rate"], u["total"],
                           u["success"], u["total"] - u["success"]],
                   alignments=[_LEFT, _RIGHT, _RIGHT, _RIGHT, _RIGHT],
                   number_formats=["", "0.0", "0", "0", "0"])
    _autosize(ws, min_width=18)


def _add_long_monthly_sheet(wb, data):
    """One row per (brand, month) - pivot-friendly in Excel / Power BI / Sheets."""
    ws = wb.create_sheet("Monthly trends (long)")
    ws["A1"] = ("Monthly trends — long format "
                "(dung de ve bieu do trong Excel / Power BI)")
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:I1")
    headers = ["Brand", "Thang", "Total", "Rated", "Response (%)",
               "Order (%)", "Perfect 10 (%)", "CSAT Score (/10)", "Source"]
    _write_header_row(ws, 3, headers)

    rates = data.get("rates", {})
    csat_brand = data.get("csat_by_brand", {})

    by_brand_month = {"BON": {m["month"]: m for m in rates.get("BON", [])},
                      "ORD": {m["month"]: m for m in rates.get("ORD", [])}}
    csat_by_month = {"BON": {m["month"]: m for m in csat_brand.get("BON", {}).get("monthly", [])},
                     "ORD": {m["month"]: m for m in csat_brand.get("ORD", {}).get("monthly", [])}}
    months = rates.get("months", [])
    row = 4
    for brand in ("BON", "ORD"):
        for m in months:
            r = by_brand_month[brand].get(m, {})
            c = csat_by_month[brand].get(m, {})
            _write_row(ws, row, [
                brand, m,
                r.get("total", 0), r.get("rated", 0),
                r.get("response_rate", 0), r.get("order_rate", 0),
                c.get("perfect_rate", 0), c.get("score", 0),
                "rates + csat_by_brand",
            ], alignments=[_CENTER, _CENTER] + [_RIGHT] * 6 + [_LEFT],
               number_formats=["", "", "0", "0", "0.0", "0.0", "0.0", "0.0", ""])
            row += 1
    ws.freeze_panes = "A4"
    _autosize(ws, min_width=11, max_width=20)


def build_export_workbook(data):
    """Build the multi-sheet Excel report from the same payload the dashboard renders."""
    wb = Workbook()
    wb.remove(wb.active)

    _add_overview_sheet(wb, data)
    _add_complaint_sheet(wb, data)
    _add_warranty_sheet(wb, data)
    _add_csat_sheets(wb, data)
    _add_purchase_sheets(wb, data)
    _add_ops_sheets(wb, data)
    _add_long_monthly_sheet(wb, data)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@app.route("/api/export")
def api_export():
    """Generate the Excel report using the same filters as /api/kpis."""
    team = request.args.get("team", "all")
    mode = request.args.get("mode", "range")
    months = int(request.args.get("months", "1"))
    month = request.args.get("date", "")

    from_date, to_date, mode, month = resolve_date_range(mode, month, months)
    team_ids_map = {"ord": [14], "bon": [15], "all": TEAM_IDS}
    team_ids = team_ids_map.get(team, TEAM_IDS)

    tickets = fetch_tickets(team_ids, from_date, to_date)
    delivery_orders = fetch_delivery_orders(from_date, to_date)

    payload = {
        "team": team,
        "mode": mode,
        "month": month,
        "period": f"{from_date[:10]} → {to_date[:10]}",
        "complaints": compute_complaint_kpis(tickets, from_date, to_date),
        "retention": compute_retention_rate(tickets),
        "warranty": compute_warranty_kpis(tickets, from_date, to_date),
        "csat": compute_csat(tickets),
        "csat_by_brand": compute_csat_by_brand(tickets, to_date),
        "satisfaction": compute_satisfaction_split(tickets),
        "rates": compute_response_and_order_rate(tickets, to_date),
        "sla": compute_sla(tickets),
        "purchase": compute_purchase_metrics(months),
        "delivery": compute_delivery_metrics(delivery_orders),
        "bd_progress": compute_bd_progress_metrics(delivery_orders),
    }

    xlsx_bytes = build_export_workbook(payload)

    team_label = {"all": "all", "ord": "ord", "bon": "bon"}.get(team, team)
    filename = (f"cs-dashboard_{team_label}_"
                f"{from_date[:10].replace('-','')}_{to_date[:10].replace('-','')}.xlsx")

    return send_file(
        BytesIO(xlsx_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── PDF report (WeasyPrint) ──────────────────────────────────────────────────
# A4 portrait, multi-section report with cover, summary KPIs, and one page
# per dashboard section. Charts are sent from the frontend as base64 PNGs and
# embedded inline (WeasyPrint handles data: URLs).

_TEAM_LABELS = {"all": "Tat ca team (Bonario + Ordinaire)",
                "ord": "Khach Ordinaire",
                "bon": "Khach Bonario"}
_MONTH_LABELS = {1: "1 thang qua", 3: "3 thang qua", 6: "6 thang qua", 12: "12 thang qua"}


def _build_report_context(data: dict, team: str, months: int) -> dict:
    """Shape the KPI payload for the report template."""
    return {
        "data": {
            **data,
            "team_label": _TEAM_LABELS.get(team, team),
            "month_label": _MONTH_LABELS.get(months, f"{months} thang qua"),
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "totals": {"tickets": data.get("csat", {}).get("total", 0)},
        }
    }


@app.route("/api/export.pdf", methods=["POST"])
def api_export_pdf():
    """Render the report template to PDF using WeasyPrint.

    Body: JSON with `filters` (same query shape as /api/kpis) and `chartImages`
    (dict of canvasId -> base64 PNG data URL).
    """
    body = request.get_json(silent=True) or {}
    filters = body.get("filters", {}) or {}
    chart_images = body.get("chartImages", {}) or {}

    team = filters.get("team", "all")
    mode = filters.get("mode", "range")
    months = int(filters.get("months", 1))
    month = filters.get("date", "") or filters.get("month", "")

    from_date, to_date, mode, month = resolve_date_range(mode, month, months)
    team_ids_map = {"ord": [14], "bon": [15], "all": TEAM_IDS}
    team_ids = team_ids_map.get(team, TEAM_IDS)

    tickets = fetch_tickets(team_ids, from_date, to_date)
    delivery_orders = fetch_delivery_orders(from_date, to_date)

    payload = {
        "team": team,
        "mode": mode,
        "month": month,
        "period": f"{from_date[:10]} → {to_date[:10]}",
        "complaints": compute_complaint_kpis(tickets, from_date, to_date),
        "retention": compute_retention_rate(tickets),
        "warranty": compute_warranty_kpis(tickets, from_date, to_date),
        "csat": compute_csat(tickets),
        "csat_by_brand": compute_csat_by_brand(tickets, to_date),
        "satisfaction": compute_satisfaction_split(tickets),
        "rates": compute_response_and_order_rate(tickets, to_date),
        "sla": compute_sla(tickets),
        "purchase": compute_purchase_metrics(months),
        "delivery": compute_delivery_metrics(delivery_orders),
        "bd_progress": compute_bd_progress_metrics(delivery_orders),
    }
    ctx = _build_report_context(payload, team, months)
    ctx["data"]["charts"] = chart_images

    # Lazy import — WeasyPrint takes ~150ms to load.
    from weasyprint import HTML
    html_str = render_template("report.html", **ctx)
    base_url = str(BASE_DIR / "templates")
    pdf_bytes = HTML(string=html_str, base_url=base_url).write_pdf()

    team_label = {"all": "all", "ord": "ord", "bon": "bon"}.get(team, team)
    filename = (f"cs-dashboard_{team_label}_"
                f"{from_date[:10].replace('-','')}_{to_date[:10].replace('-','')}.pdf")

    return send_file(
        BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("CS Dashboard — http://localhost:5100")
    app.run(host="0.0.0.0", port=5100, debug=True)
